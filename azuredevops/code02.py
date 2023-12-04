import requests
import openpyxl

# Detalhes de autenticação e organização do Azure DevOps
organization = "MLTECNO02"
personal_access_token = "gd4ic4tssawj4rnrmtbgs4p56npg25rde3p2cshen4kabmctuaga"
project_id = "cloudengennier"
item_backlog_pai_id = "Codificar Network"

# Nome do arquivo Excel
nome_planilha = 'conta01.xlsx'
caminho_arquivo = f"F:/Codificação/azuredevops/{nome_planilha}"

# Carrega o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo)
sheet = workbook.active

# Lista para armazenar os dados da segunda coluna
dados_segunda_coluna = []

# Itera sobre as células da segunda coluna e armazena os dados na lista
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=2).value  # Acessa a segunda coluna (column=2)
    dados_segunda_coluna.append(cell_value)

# URL do endpoint para criar tarefas
url = f"https://dev.azure.com/{organization}/{project_id}/_apis/wit/workitems/$Task?api-version=6.0"

# Cabeçalho com o token de acesso e tipo de conteúdo
headers = {
    "Authorization": f"Basic {personal_access_token}",
    "Content-Type": "application/json"
}

# Itera sobre os valores da segunda coluna para criar tarefas no Azure DevOps
for valor in dados_segunda_coluna:
    # Corpo da requisição para criar uma nova tarefa e relacioná-la ao item de backlog pai
    data = [
        {
            "op": "add",
            "path": "/fields/System.Title",
            "value": valor
        },
        {
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": f"https://dev.azure.com/{organization}/{project_id}/_apis/wit/workItems/{item_backlog_pai_id}"
            }
        }
    ]

    # Faz a requisição POST para criar a tarefa dentro do item de backlog pai
    response = requests.post(url, headers=headers, json=data)

    if response.status_code == 200:
        print(f"Tarefa '{valor}' criada com sucesso dentro do item de backlog específico.")
    else:
        print(f"Erro ao criar tarefa '{valor}' - Código: {response.status_code} - Mensagem: {response.text}")
