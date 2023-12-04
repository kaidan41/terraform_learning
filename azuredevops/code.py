import requests
import base64
import openpyxl


#Todas as barras do caminho devem ser invertidas pro Phyton efetuar a leitura

NomeDaPlanilha = 'conta01.xlsx'

# Caminho do arquivo Excel
caminho_arquivo = f"F:/Codificação/azuredevops/{NomeDaPlanilha}"


# Configurações
organization_name = "MLTECNO02"
personal_access_token = "gd4ic4tssawj4rnrmtbgs4p56npg25rde3p2cshen4kabmctuaga"
project_name = "cloudengennier"  # Nome do projeto que você quer selecionar

# ID do item de backlog pai onde você deseja criar a tarefa
item_backlog_pai_id = "Codificar Network"

# Endpoint para listar projetos
url = f"https://dev.azure.com/{organization_name}/_apis/projects/{project_name}?api-version=6.0"

# Cabeçalho com o token de acesso
headers = {
    "Authorization": f"Basic {base64.b64encode(f':{personal_access_token}'.encode()).decode()}"
}

# Faz a requisição GET para obter detalhes do projeto específico
response = requests.get(url, headers=headers)

if response.status_code == 200:
    project_details = response.json()
    print("")
    print(f"Projeto selecionado é o : '{project_name}':")
    print("")
    print("Detalhes do projeto")
    print(project_details)  # Aqui você pode processar os detalhes como desejar
else:
    print(f"Erro: {response.status_code} - {response.text}")


#Lendo os dados da Planilha em Excell


# Carrega o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo)

# Seleciona a planilha desejada
sheet = workbook.active

# Lista para armazenar os dados da segunda coluna
dados_segunda_coluna = []

# Itera sobre as células da segunda coluna e armazena os dados na lista
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=2).value  # Acessa a segunda coluna (column=2)
    dados_segunda_coluna.append(cell_value)

# Imprime os dados armazenados na lista, um por linha
for valor in dados_segunda_coluna:
    print(valor)

print("")

print("Iremos começar a criar as tasks no azure Devops Aguarde por favor ....")

# URL do endpoint para criar tarefas
url = f"https://dev.azure.com/{organization_name}/{project_name}/_apis/wit/workitems/$Task?api-version=6.0"



# Cabeçalho com o token de acesso e tipo de conteúdo
headers = {
    "Authorization": f"Basic {personal_access_token}",
    "Content-Type": "application/json"
}

# Itera sobre os valores da segunda coluna para criar tarefas no Azure DevOps
for valor in dados_segunda_coluna:
    # Corpo da requisição para criar uma nova tarefa e relacioná-la ao item de backlog pai
    data = [
        {
            "op": "add",
            "path": "/fields/System.Title",
            "value": valor
        },
        {
            "op": "add",
            "path": "/relations/-",
            "value": {
                "rel": "System.LinkTypes.Hierarchy-Reverse",
                "url": f"https://dev.azure.com/{organization_name}/{project_name}/_apis/wit/workItems/{item_backlog_pai_id}"
            }
        }
    ]

    # Faz a requisição POST para criar a tarefa dentro do item de backlog pai
    response = requests.post(url, headers=headers, json=data)

    if response.status_code == 200:
        print(f"Tarefa '{valor}' criada com sucesso dentro do item de backlog específico.")
    else:
        print(f"Erro ao criar tarefa '{valor}' - Código: {response.status_code} - Mensagem: {response.text}")
